from django.contrib import admin
from django.utils.html import format_html
from django.urls import reverse, path
from django.utils import timezone
from django.shortcuts import render, redirect
from django.contrib import messages
from django import forms
from .models import Medicine, MedicineTransaction, StockRequest
import pandas as pd


class ExcelImportForm(forms.Form):
    excel_file = forms.FileField(
        label='Select Excel File',
        help_text='Upload an Excel file (.xlsx or .xls) with medicine data'
    )


@admin.register(Medicine)
class MedicineAdmin(admin.ModelAdmin):
    list_display = [
        'name', 'category', 'current_stock', 'minimum_stock_level', 
        'stock_status_colored', 'unit_price', 'expiry_date', 'is_active'
    ]
    list_filter = [
        'category', 'manufacturer', 'is_active', 'expiry_date'
    ]
    search_fields = ['name', 'generic_name', 'manufacturer', 'description']
    ordering = ['name']
    # Make current_stock readonly - stock changes must go through MedicineTransaction
    readonly_fields = ['current_stock', 'created_at', 'updated_at', 'stock_status', 'is_low_stock']
    
    fieldsets = (
        ('Basic Information', {
            'fields': ('name', 'generic_name', 'category', 'manufacturer', 'description')
        }),
        ('Stock Management', {
            'fields': ('current_stock', 'minimum_stock_level', 'unit', 'unit_price'),
            'description': '⚠️ Current stock is READ-ONLY. To change stock, create a Medicine Transaction (Received/Issued/Adjustment) below.'
        }),
        ('Batch Information', {
            'fields': ('batch_number', 'expiry_date'),
            'classes': ('collapse',)
        }),
        ('Status & Timestamps', {
            'fields': ('is_active', 'stock_status', 'is_low_stock', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def stock_status_colored(self, obj):
        """Display stock status with colors"""
        status = obj.stock_status
        colors = {
            'out_of_stock': 'red',
            'low_stock': 'orange',
            'adequate': 'green'
        }
        return format_html(
            '<span style="color: {}; font-weight: bold;">{}</span>',
            colors.get(status, 'black'),
            status.replace('_', ' ').title()
        )
    stock_status_colored.short_description = 'Stock Status'
    stock_status_colored.admin_order_field = 'current_stock'
    
    actions = ['mark_as_low_stock_alert', 'deactivate_medicines']
    
    def mark_as_low_stock_alert(self, request, queryset):
        """Create stock requests for selected medicines with low stock"""
        count = 0
        for medicine in queryset:
            if medicine.is_low_stock:
                StockRequest.objects.get_or_create(
                    medicine=medicine,
                    status='pending',
                    defaults={
                        'requested_quantity': medicine.minimum_stock_level * 2,
                        'current_stock': medicine.current_stock,
                        'reason': f'Auto-generated: Low stock alert for {medicine.name}',
                        'priority': 'high',
                        'requested_by': request.user
                    }
                )
                count += 1
        self.message_user(request, f'{count} stock requests created for low stock medicines.')
    mark_as_low_stock_alert.short_description = 'Create stock requests for low stock items'
    
    def deactivate_medicines(self, request, queryset):
        """Deactivate selected medicines"""
        updated = queryset.update(is_active=False)
        self.message_user(request, f'{updated} medicines deactivated.')
    deactivate_medicines.short_description = 'Deactivate selected medicines'
    
    def get_urls(self):
        """Add custom URLs for Excel import"""
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel), name='medicines_medicine_import'),
            path('download-template/', self.admin_site.admin_view(self.download_template), name='medicines_medicine_template'),
        ]
        return custom_urls + urls
    
    def import_excel(self, request):
        """Handle Excel file upload and import"""
        if request.method == 'POST':
            form = ExcelImportForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES['excel_file']
                
                # Validate file extension
                if not excel_file.name.endswith(('.xlsx', '.xls')):
                    messages.error(request, 'Invalid file format. Please upload an Excel file (.xlsx or .xls)')
                    return redirect('..')
                
                try:
                    # Read Excel file
                    df = pd.read_excel(excel_file)
                    
                    # Required columns
                    required_columns = ['name', 'category', 'manufacturer', 'description', 'unit', 'unit_price', 'batch_number']
                    missing_columns = [col for col in required_columns if col not in df.columns]
                    
                    if missing_columns:
                        messages.error(request, f'Missing required columns: {", ".join(missing_columns)}')
                        return redirect('..')
                    
                    # Optional columns with defaults
                    if 'generic_name' not in df.columns:
                        df['generic_name'] = ''
                    if 'minimum_stock_level' not in df.columns:
                        df['minimum_stock_level'] = 10
                    
                    # Validate category values
                    valid_categories = ['tablet', 'capsule', 'syrup', 'injection', 'ointment', 'drops', 'other']
                    
                    created_count = 0
                    skipped_count = 0
                    errors = []
                    
                    for index, row in df.iterrows():
                        try:
                            # Validate required fields
                            name = str(row['name']).strip()
                            if not name or pd.isna(row['name']):
                                errors.append(f"Row {index + 2}: Medicine name is required")
                                skipped_count += 1
                                continue
                            
                            # Validate category
                            category = str(row['category']).lower().strip()
                            if category not in valid_categories:
                                errors.append(f"Row {index + 2}: Invalid category '{row['category']}'. Must be one of: {', '.join(valid_categories)}")
                                skipped_count += 1
                                continue
                            
                            # Validate other required fields
                            if pd.isna(row['manufacturer']) or not str(row['manufacturer']).strip():
                                errors.append(f"Row {index + 2}: Manufacturer is required")
                                skipped_count += 1
                                continue
                            
                            if pd.isna(row['description']) or not str(row['description']).strip():
                                errors.append(f"Row {index + 2}: Description is required")
                                skipped_count += 1
                                continue
                            
                            if pd.isna(row['unit']) or not str(row['unit']).strip():
                                errors.append(f"Row {index + 2}: Unit is required")
                                skipped_count += 1
                                continue
                            
                            if pd.isna(row['unit_price']):
                                errors.append(f"Row {index + 2}: Unit price is required")
                                skipped_count += 1
                                continue
                            
                            if pd.isna(row['batch_number']) or not str(row['batch_number']).strip():
                                errors.append(f"Row {index + 2}: Batch number is required")
                                skipped_count += 1
                                continue
                            
                            # Check if medicine already exists
                            if Medicine.objects.filter(name__iexact=name).exists():
                                errors.append(f"Row {index + 2}: Medicine '{name}' already exists")
                                skipped_count += 1
                                continue
                            
                            # Create medicine with current_stock = 0
                            Medicine.objects.create(
                                name=name,
                                generic_name=str(row['generic_name']).strip() if pd.notna(row['generic_name']) else '',
                                category=category,
                                manufacturer=str(row['manufacturer']).strip(),
                                description=str(row['description']).strip(),
                                current_stock=0,  # Always start with 0
                                minimum_stock_level=int(row['minimum_stock_level']) if pd.notna(row['minimum_stock_level']) else 10,
                                unit=str(row['unit']).strip(),
                                unit_price=float(row['unit_price']),
                                batch_number=str(row['batch_number']).strip(),
                                is_active=True
                            )
                            created_count += 1
                            
                        except Exception as e:
                            errors.append(f"Row {index + 2}: {str(e)}")
                            skipped_count += 1
                    
                    # Display results
                    if created_count > 0:
                        messages.success(request, f'Successfully imported {created_count} medicines')
                    if skipped_count > 0:
                        messages.warning(request, f'{skipped_count} rows were skipped')
                    if errors:
                        for error in errors[:10]:  # Show first 10 errors
                            messages.error(request, error)
                        if len(errors) > 10:
                            messages.error(request, f'... and {len(errors) - 10} more errors')
                    
                    return redirect('..')
                    
                except Exception as e:
                    messages.error(request, f'Failed to process Excel file: {str(e)}')
                    return redirect('..')
        else:
            form = ExcelImportForm()
        
        context = {
            'form': form,
            'title': 'Import Medicines from Excel',
            'site_header': 'Medicine Import',
            'site_title': 'Import',
            'opts': self.model._meta,
        }
        return render(request, 'admin/medicines/import_excel.html', context)
    
    def download_template(self, request):
        """Download Excel template for medicine import"""
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Medicines Template"
        
        # Headers
        headers = [
            'name', 'generic_name', 'category', 'manufacturer', 
            'description', 'minimum_stock_level', 'unit', 'unit_price', 'batch_number'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Add sample data
        sample_data = [
            ['Paracetamol 500mg', 'Paracetamol', 'tablet', 'ABC Pharma', 'Pain reliever and fever reducer', 50, 'pieces', 2.50, 'BATCH001'],
            ['Amoxicillin 250mg', 'Amoxicillin', 'capsule', 'XYZ Labs', 'Antibiotic for bacterial infections', 30, 'pieces', 5.00, 'BATCH002'],
            ['Cough Syrup', 'Dextromethorphan', 'syrup', 'MediCare', 'For dry cough relief', 20, 'bottles', 75.00, 'BATCH003']
        ]
        
        for row_num, data in enumerate(sample_data, 2):
            for col_num, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions")
        instructions = [
            ["Medicine Import Template Instructions"],
            [""],
            ["Required Columns:"],
            ["- name: Medicine name (must be unique)"],
            ["- category: Must be one of: tablet, capsule, syrup, injection, ointment, drops, other"],
            ["- manufacturer: Manufacturer name"],
            ["- description: Medicine description"],
            ["- unit: Unit of measurement (e.g., pieces, tablets, bottles, ml)"],
            ["- unit_price: Price per unit (numeric value)"],
            ["- batch_number: Batch identification number"],
            [""],
            ["Optional Columns (with defaults):"],
            ["- generic_name: Generic/chemical name (default: empty)"],
            ["- minimum_stock_level: Minimum quantity alert (default: 10)"],
            [""],
            ["Important Notes:"],
            ["- All medicines will be imported with current_stock = 0"],
            ["- Pharmacists can then request stock for these medicines"],
            ["- Duplicate medicine names will be skipped"],
            ["- All required fields must have values or row will be skipped"],
            ["- Invalid category values will be rejected"]
        ]
        
        for row_num, instruction in enumerate(instructions, 1):
            cell = ws_instructions.cell(row=row_num, column=1, value=instruction[0])
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif "Required Columns:" in instruction[0] or "Optional Columns" in instruction[0] or "Important Notes:" in instruction[0]:
                cell.font = Font(bold=True)
        
        ws_instructions.column_dimensions['A'].width = 80
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=medicine_import_template.xlsx'
        wb.save(response)
        
        return response
    
    def changelist_view(self, request, extra_context=None):
        """Add import button to changelist view"""
        extra_context = extra_context or {}
        extra_context['show_import_button'] = True
        return super().changelist_view(request, extra_context=extra_context)


@admin.register(MedicineTransaction)
class MedicineTransactionAdmin(admin.ModelAdmin):
    list_display = [
        'medicine', 'transaction_type', 'quantity', 'date', 
        'performed_by', 'supplier', 'patient_display'
    ]
    list_filter = [
        'transaction_type', 'date', 'performed_by', 'medicine__category'
    ]
    search_fields = [
        'medicine__name', 'supplier', 'patient', 'patient_record__name', 
        'patient_record__employee_student_id', 'reference_number', 'remarks'
    ]
    ordering = ['-date']
    readonly_fields = ['created_at', 'patient_display']
    date_hierarchy = 'date'
    
    fieldsets = (
        ('Transaction Details', {
            'fields': ('medicine', 'transaction_type', 'quantity', 'date')
        }),
        ('Reference Information', {
            'fields': ('reference_number', 'supplier', 'patient_record', 'patient_display'),
            'description': 'Supplier for received items, Patient for issued items'
        }),
        ('System Information', {
            'fields': ('performed_by', 'remarks', 'created_at'),
            'classes': ('collapse',)
        }),
    )
    
    def patient_display(self, obj):
        """Display patient name from patient_record or fallback to patient field"""
        if obj.patient_record:
            return f"{obj.patient_record.name} ({obj.patient_record.employee_student_id})"
        return obj.patient or 'N/A'
    patient_display.short_description = 'Patient'
    
    def get_readonly_fields(self, request, obj=None):
        """Make performed_by readonly when editing"""
        readonly = list(self.readonly_fields)
        if obj:  # Editing existing object
            readonly.extend(['medicine', 'transaction_type', 'quantity', 'date'])
        return readonly
    
    def save_model(self, request, obj, form, change):
        """Auto-set performed_by for new transactions and adjust medicine stock"""
        if not change:  # Creating new object
            obj.performed_by = request.user
        super().save_model(request, obj, form, change)

        # Adjust medicine stock after saving transaction
        medicine = obj.medicine
        if obj.transaction_type == 'received':
            medicine.current_stock += obj.quantity
        elif obj.transaction_type == 'issued':
            medicine.current_stock = max(0, medicine.current_stock - obj.quantity)
        elif obj.transaction_type in ['expired', 'adjustment']:
            medicine.current_stock = max(0, medicine.current_stock - obj.quantity)
        medicine.save(skip_stock_validation=True)


@admin.register(StockRequest)
class StockRequestAdmin(admin.ModelAdmin):
    list_display = [
        'medicine', 'requested_quantity', 'priority', 'status', 
        'requested_by', 'requested_date', 'approved_by'
    ]
    list_filter = [
        'status', 'priority', 'requested_date', 'approved_date', 'medicine__category'
    ]
    search_fields = [
        'medicine__name', 'reason', 'notes', 'requested_by__username'
    ]
    ordering = ['-requested_date']
    readonly_fields = [
        'requested_by', 'requested_date', 'approved_by', 'approved_date', 'current_stock'
    ]
    date_hierarchy = 'requested_date'
    
    fieldsets = (
        ('Request Details', {
            'fields': ('medicine', 'requested_quantity', 'current_stock', 'priority')
        }),
        ('Justification', {
            'fields': ('reason', 'estimated_usage_days')
        }),
        ('Status Management', {
            'fields': ('status', 'expected_delivery_date', 'notes'),
            'description': '⚠️ IMPORTANT: When approving via API, expiry_date and batch_number must be provided. Admin panel approvals should be followed by updating the medicine\'s expiry date manually.'
        }),
        ('System Information', {
            'fields': ('requested_by', 'requested_date', 'approved_by', 'approved_date'),
            'classes': ('collapse',)
        }),
    )
    
    def get_readonly_fields(self, request, obj=None):
        """Dynamic readonly fields based on user permissions and status"""
        readonly = list(self.readonly_fields)
        
        if obj and obj.status in ['approved', 'rejected', 'received']:
            # Can't edit approved/rejected/received requests
            readonly.extend(['requested_quantity', 'priority', 'reason'])
        
        return readonly
    
    def save_model(self, request, obj, form, change):
        """Auto-set request details and handle status changes"""
        if not change:  # Creating new object
            obj.requested_by = request.user
            obj.current_stock = obj.medicine.current_stock
        else:  # Updating existing object
            # Handle status changes
            if 'status' in form.changed_data:
                if obj.status in ['approved', 'rejected'] and not obj.approved_by:
                    obj.approved_by = request.user
                    obj.approved_date = timezone.now()
        
        super().save_model(request, obj, form, change)
    
    actions = ['approve_requests', 'mark_as_ordered']
    
    def approve_requests(self, request, queryset):
        """Approve selected pending requests - Note: Requires manual expiry date update"""
        pending_requests = queryset.filter(status='pending')
        if not pending_requests.exists():
            self.message_user(request, 'No pending requests selected.', level='warning')
            return
        
        updated = pending_requests.update(
            status='approved',
            approved_by=request.user,
            approved_date=timezone.now()
        )
        
        self.message_user(
            request, 
            f'{updated} stock requests approved. ⚠️ REMEMBER: Update expiry_date and batch_number for each medicine manually or use the API endpoint.',
            level='warning'
        )
    approve_requests.short_description = '✓ Approve selected pending requests (requires manual expiry update)'
    
    def mark_as_ordered(self, request, queryset):
        """Mark approved requests as ordered"""
        approved_requests = queryset.filter(status='approved')
        updated = approved_requests.update(status='ordered')
        self.message_user(request, f'{updated} stock requests marked as ordered.')
    mark_as_ordered.short_description = 'Mark approved requests as ordered'
