from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import F, Q, Sum, Count
from django.utils import timezone
from datetime import timedelta
import pandas as pd
import openpyxl
from io import BytesIO
from .models import Medicine, MedicineTransaction, StockRequest
from .serializers import (
    MedicineSerializer, MedicineCreateUpdateSerializer, PatientMedicineSerializer,
    MedicineTransactionSerializer, StockRequestSerializer, 
    StockRequestApprovalSerializer
)
from users.permissions import IsMedicalStaff, IsPharmacist, IsDoctorOrAdmin, IsPharmacistOrPatientReadOnly


class MedicineViewSet(viewsets.ModelViewSet):
    """ViewSet for managing medicines"""
    queryset = Medicine.objects.all()
    serializer_class = MedicineSerializer
    permission_classes = [IsAuthenticated, IsPharmacistOrPatientReadOnly]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['category', 'manufacturer', 'is_active']
    search_fields = ['name', 'generic_name', 'description', 'batch_number']
    
    def get_serializer_class(self):
        """Use different serializers for different actions"""
        if self.action in ['create', 'update', 'partial_update']:
            return MedicineCreateUpdateSerializer
        return MedicineSerializer

    def get_queryset(self):
        """Filter medicines based on user type and query parameters"""
        queryset = super().get_queryset()
        user_type = getattr(self.request.user, 'user_type', None)
        if user_type in ['student', 'employee']:
            queryset = queryset.filter(is_active=True)

        expiry_status = self.request.query_params.get('expiry_status')
        if expiry_status == 'expired':
            queryset = queryset.filter(expiry_date__lt=timezone.now().date())
        elif expiry_status == 'expiring_soon':
            thirty_days_later = timezone.now().date() + timedelta(days=30)
            queryset = queryset.filter(
                expiry_date__lte=thirty_days_later,
                expiry_date__gte=timezone.now().date()
            )

        stock_status = self.request.query_params.get('stock_status')
        if stock_status == 'low_stock':
            queryset = queryset.filter(current_stock__lte=F('minimum_stock_level'))
        elif stock_status == 'out_of_stock':
            queryset = queryset.filter(current_stock=0)
        elif stock_status == 'adequate':
            queryset = queryset.filter(current_stock__gt=F('minimum_stock_level'))

        return queryset
    
    @action(detail=False, methods=['get'])
    def low_stock(self, request):
        """Get medicines with low stock"""
        medicines = Medicine.objects.filter(
            current_stock__lte=F('minimum_stock_level'),
            is_active=True
        )
        serializer = self.get_serializer(medicines, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def out_of_stock(self, request):
        """Get medicines that are out of stock"""
        medicines = Medicine.objects.filter(current_stock=0, is_active=True)
        serializer = self.get_serializer(medicines, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def expiring_soon(self, request):
        """Get medicines expiring in next 30 days"""
        from datetime import timedelta
        thirty_days_later = timezone.now().date() + timedelta(days=30)
        medicines = Medicine.objects.filter(
            expiry_date__lte=thirty_days_later,
            expiry_date__gte=timezone.now().date(),
            is_active=True
        )
        serializer = self.get_serializer(medicines, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'], permission_classes=[IsAuthenticated])
    def import_excel(self, request):
        """Import medicines from Excel file (Admin only)"""
        # Check if user is admin
        if request.user.user_type != 'admin':
            return Response(
                {'error': 'Only administrators can import medicines'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        file = request.FILES.get('file')
        if not file:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate file extension
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Read Excel file
            df = pd.read_excel(file)
            
            # Required columns
            required_columns = ['name', 'category', 'manufacturer', 'description', 'unit', 'unit_price', 'batch_number']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return Response(
                    {'error': f'Missing required columns: {", ".join(missing_columns)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Optional columns with defaults
            optional_columns = {
                'generic_name': '',
                'minimum_stock_level': 10
            }
            
            # Add missing optional columns with defaults
            for col, default_value in optional_columns.items():
                if col not in df.columns:
                    df[col] = default_value
            
            # Validate category values
            valid_categories = ['tablet', 'capsule', 'syrup', 'injection', 'ointment', 'drops', 'other']
            
            created_count = 0
            skipped_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Validate required fields
                    name = str(row['name']).strip()
                    if not name or pd.isna(row['name']):
                        errors.append(f"Row {index + 2}: Medicine name is required")
                        skipped_count += 1
                        continue
                    
                    # Validate category
                    category = str(row['category']).lower().strip()
                    if category not in valid_categories:
                        errors.append(f"Row {index + 2}: Invalid category '{row['category']}'. Must be one of: {', '.join(valid_categories)}")
                        skipped_count += 1
                        continue
                    
                    # Validate other required fields
                    if pd.isna(row['manufacturer']) or not str(row['manufacturer']).strip():
                        errors.append(f"Row {index + 2}: Manufacturer is required")
                        skipped_count += 1
                        continue
                    
                    if pd.isna(row['description']) or not str(row['description']).strip():
                        errors.append(f"Row {index + 2}: Description is required")
                        skipped_count += 1
                        continue
                    
                    if pd.isna(row['unit']) or not str(row['unit']).strip():
                        errors.append(f"Row {index + 2}: Unit is required")
                        skipped_count += 1
                        continue
                    
                    if pd.isna(row['unit_price']):
                        errors.append(f"Row {index + 2}: Unit price is required")
                        skipped_count += 1
                        continue
                    
                    if pd.isna(row['batch_number']) or not str(row['batch_number']).strip():
                        errors.append(f"Row {index + 2}: Batch number is required")
                        skipped_count += 1
                        continue
                    
                    # Check if medicine already exists
                    if Medicine.objects.filter(name__iexact=name).exists():
                        errors.append(f"Row {index + 2}: Medicine '{name}' already exists")
                        skipped_count += 1
                        continue
                    
                    # Create medicine with current_stock = 0
                    Medicine.objects.create(
                        name=name,
                        generic_name=str(row['generic_name']).strip() if pd.notna(row['generic_name']) else '',
                        category=category,
                        manufacturer=str(row['manufacturer']).strip(),
                        description=str(row['description']).strip(),
                        current_stock=0,  # Always start with 0
                        minimum_stock_level=int(row['minimum_stock_level']) if pd.notna(row['minimum_stock_level']) else 10,
                        unit=str(row['unit']).strip(),
                        unit_price=float(row['unit_price']),
                        batch_number=str(row['batch_number']).strip(),
                        is_active=True
                    )
                    created_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {index + 2}: {str(e)}")
                    skipped_count += 1
            
            return Response({
                'success': True,
                'message': f'Import completed. Created: {created_count}, Skipped: {skipped_count}',
                'created': created_count,
                'skipped': skipped_count,
                'errors': errors if errors else None
            })
            
        except Exception as e:
            return Response(
                {'error': f'Failed to process Excel file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def download_template(self, request):
        """Download Excel template for medicine import"""
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Medicines Template"
        
        # Headers
        headers = [
            'name', 'generic_name', 'category', 'manufacturer', 
            'description', 'minimum_stock_level', 'unit', 'unit_price', 'batch_number'
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Add sample data
        sample_data = [
            ['Paracetamol 500mg', 'Paracetamol', 'tablet', 'ABC Pharma', 'Pain reliever', 50, 'pieces', 2.50, 'BATCH001'],
            ['Amoxicillin 250mg', 'Amoxicillin', 'capsule', 'XYZ Labs', 'Antibiotic', 30, 'pieces', 5.00, 'BATCH002'],
            ['Cough Syrup', 'Dextromethorphan', 'syrup', 'MediCare', 'For dry cough', 20, 'bottles', 75.00, 'BATCH003']
        ]
        
        for row_num, data in enumerate(sample_data, 2):
            for col_num, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions")
        instructions = [
            ["Medicine Import Template Instructions"],
            [""],
            ["Required Columns:"],
            ["- name: Medicine name (must be unique)"],
            ["- category: Must be one of: tablet, capsule, syrup, injection, ointment, drops, other"],
            ["- manufacturer: Manufacturer name"],
            ["- description: Medicine description"],
            ["- unit: Unit of measurement (e.g., pieces, tablets, bottles, ml)"],
            ["- unit_price: Price per unit (numeric value)"],
            ["- batch_number: Batch identification number"],
            [""],
            ["Optional Columns (with defaults):"],
            ["- generic_name: Generic/chemical name (default: empty)"],
            ["- minimum_stock_level: Minimum quantity alert (default: 10)"],
            [""],
            ["Important Notes:"],
            ["- All medicines will be imported with current_stock = 0"],
            ["- Pharmacists can then request stock for these medicines"],
            ["- Duplicate medicine names will be skipped"],
            ["- All required fields must have values or row will be skipped"],
            ["- Invalid category values will be rejected"]
        ]
        
        for row_num, instruction in enumerate(instructions, 1):
            ws_instructions.cell(row=row_num, column=1, value=instruction[0])
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=medicine_import_template.xlsx'
        wb.save(response)
        
        return response
    
    @action(detail=True, methods=['post'])
    def adjust_stock(self, request, pk=None):
        """Manually adjust medicine stock"""
        medicine = self.get_object()
        new_stock = request.data.get('new_stock')
        reason = request.data.get('reason', 'Manual adjustment')
        
        if new_stock is None:
            return Response({'error': 'new_stock is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            new_stock = int(new_stock)
            if new_stock < 0:
                return Response({'error': 'Stock cannot be negative'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Create adjustment transaction
            quantity_diff = abs(medicine.current_stock - new_stock)
            transaction_type = 'received' if new_stock > medicine.current_stock else 'adjustment'
            
            MedicineTransaction.objects.create(
                medicine=medicine,
                transaction_type=transaction_type,
                quantity=quantity_diff,
                date=timezone.now(),
                performed_by=request.user,
                remarks=reason
            )
            
            medicine.current_stock = new_stock
            # Use skip_stock_validation to bypass direct stock change validation
            medicine.save(skip_stock_validation=True)
            
            return Response({
                'success': True,
                'new_stock': medicine.current_stock,
                'message': f'Stock adjusted to {new_stock}'
            })
            
        except ValueError:
            return Response({'error': 'Invalid stock value'}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def analytics(self, request):
        """Get medicine analytics and statistics"""
        total_medicines = Medicine.objects.filter(is_active=True).count()
        low_stock_count = Medicine.objects.filter(
            current_stock__lte=F('minimum_stock_level'),
            is_active=True
        ).count()
        out_of_stock_count = Medicine.objects.filter(
            current_stock=0,
            is_active=True
        ).count()
        
        # Expiry analysis
        today = timezone.now().date()
        thirty_days_later = today + timedelta(days=30)
        expired_count = Medicine.objects.filter(
            expiry_date__lt=today,
            is_active=True
        ).count()
        expiring_soon_count = Medicine.objects.filter(
            expiry_date__lte=thirty_days_later,
            expiry_date__gte=today,
            is_active=True
        ).count()
        
        # Value analysis
        total_stock_value = Medicine.objects.filter(is_active=True).aggregate(
            total_value=Sum(F('current_stock') * F('unit_price'))
        )['total_value'] or 0
        
        # Category breakdown
        category_stats = Medicine.objects.filter(is_active=True).values('category').annotate(
            count=Count('id'),
            total_stock=Sum('current_stock'),
            total_value=Sum(F('current_stock') * F('unit_price'))
        ).order_by('-count')
        
        return Response({
            'summary': {
                'total_medicines': total_medicines,
                'low_stock_count': low_stock_count,
                'out_of_stock_count': out_of_stock_count,
                'expired_count': expired_count,
                'expiring_soon_count': expiring_soon_count,
                'total_stock_value': float(total_stock_value),
                'stock_health_percentage': round(
                    ((total_medicines - low_stock_count - out_of_stock_count) / max(total_medicines, 1)) * 100, 2
                )
            },
            'categories': list(category_stats),
            'alerts': {
                'critical': out_of_stock_count + expired_count,
                'warning': low_stock_count + expiring_soon_count
            }
        })
    
    @action(detail=False, methods=['get'])
    def categories(self, request):
        """Get list of medicine categories with counts"""
        categories = Medicine.objects.filter(is_active=True).values('category').annotate(
            count=Count('id'),
            low_stock_count=Count('id', filter=Q(current_stock__lte=F('minimum_stock_level'))),
            out_of_stock_count=Count('id', filter=Q(current_stock=0))
        ).order_by('category')
        
        return Response(list(categories))


class MedicineTransactionViewSet(viewsets.ModelViewSet):
    queryset = MedicineTransaction.objects.all()
    serializer_class = MedicineTransactionSerializer
    permission_classes = [IsAuthenticated, IsPharmacist]  # Only pharmacists can access transactions
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['transaction_type', 'medicine', 'performed_by']
    search_fields = ['medicine__name', 'supplier', 'patient', 'remarks']
    ordering_fields = ['requested_date', 'approved_date', 'id']
    ordering = ['-id']
    
    def get_queryset(self):
        """Restrict access to medicine transactions. HODs can only see transactions related to their department's patients."""
        user = self.request.user
        
        # HODs can only see transactions related to their department's patients
        if getattr(user, 'user_type', None) == 'hod':
            # Filter transactions by patient_record's user department
            return MedicineTransaction.objects.filter(patient_record__user__department=user.department)
        
        # Medical staff can see all transactions
        return MedicineTransaction.objects.all()
    
    @action(detail=False, methods=['get'])
    def today(self, request):
        """Get today's transactions"""
        today = timezone.now().date()
        transactions = MedicineTransaction.objects.filter(date__date=today)
        serializer = self.get_serializer(transactions, many=True)
        return Response(serializer.data)


class StockRequestViewSet(viewsets.ModelViewSet):
    queryset = StockRequest.objects.select_related('medicine', 'requested_by', 'approved_by')
    serializer_class = StockRequestSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['status', 'priority', 'medicine', 'requested_by', 'approved_by']
    search_fields = ['medicine__name', 'reason', 'notes']
    ordering_fields = ['requested_date', 'priority', 'approved_date']
    ordering = ['-requested_date']
    
    def get_serializer_class(self):
        """Use different serializers for different actions"""
        if self.action in ['approve', 'reject']:
            return StockRequestApprovalSerializer
        return StockRequestSerializer
    
    def get_queryset(self):
        """Custom queryset with additional filtering"""
        queryset = StockRequest.objects.select_related(
            'medicine',
            'requested_by',
            'approved_by'
        )

        user = self.request.user

        # ROLE-BASED VISIBILITY
        if user.user_type == 'hod':
            queryset = queryset.filter(
                requested_by__department=user.department
            )
        elif user.user_type == 'pharmacist':
            pass  # pharmacists see all
        elif user.user_type not in ['admin', 'principal']:
            queryset = queryset.filter(requested_by=user)

        # Filter by status / priority etc (DjangoFilterBackend still works)
        days_old = self.request.query_params.get('days_old')
        if days_old:
            try:
                days_ago = timezone.now() - timedelta(days=int(days_old))
                queryset = queryset.filter(requested_date__lte=days_ago)
            except ValueError:
                pass

        if self.request.query_params.get('my_requests') == 'true':
            queryset = queryset.filter(requested_by=user)

        return queryset

    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve a stock request and increment medicine stock"""
        stock_request = self.get_object()
        if stock_request.status != 'pending':
            return Response({
                'error': 'Only pending requests can be approved'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate required fields
        expiry_date = request.data.get('expiry_date')
        if not expiry_date:
            return Response({
                'error': 'expiry_date is required when approving stock requests'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate expiry_date is not in the past
        from datetime import datetime
        try:
            expiry_date_obj = datetime.strptime(expiry_date, '%Y-%m-%d').date()
            if expiry_date_obj < timezone.now().date():
                return Response({
                    'error': 'Expiry date cannot be in the past'
                }, status=status.HTTP_400_BAD_REQUEST)
        except ValueError:
            return Response({
                'error': 'Invalid date format. Use YYYY-MM-DD'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Get optional batch_number
        batch_number = request.data.get('batch_number', '')
        
        # Update request status
        stock_request.status = 'approved'
        stock_request.approved_by = request.user
        stock_request.approved_date = timezone.now()
        if 'notes' in request.data:
            stock_request.notes = request.data['notes']
        stock_request.save()
        
        # Increment medicine stock using transaction
        medicine = stock_request.medicine
        quantity = stock_request.requested_quantity
        from django.utils import timezone as dj_timezone
        from medicines.models import MedicineTransaction
        
        MedicineTransaction.objects.create(
            medicine=medicine,
            transaction_type='received',
            quantity=quantity,
            date=dj_timezone.now(),
            performed_by=request.user,
            remarks=f"Stock request approved (ID: {stock_request.id})"
        )
        
        # Update medicine stock and expiry date
        medicine.current_stock += quantity
        medicine.expiry_date = expiry_date_obj
        if batch_number:
            medicine.batch_number = batch_number
        medicine.save(skip_stock_validation=True)
        
        return Response({
            'success': True,
            'message': 'Stock request approved and stock updated',
            'new_stock': medicine.current_stock,
            'expiry_date': medicine.expiry_date,
            'batch_number': medicine.batch_number
        })
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Reject a stock request"""
        stock_request = self.get_object()
        
        if stock_request.status != 'pending':
            return Response({
                'error': 'Only pending requests can be rejected'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        stock_request.status = 'rejected'
        stock_request.approved_by = request.user
        stock_request.approved_date = timezone.now()
        stock_request.notes = request.data.get('reason', 'Request rejected')
        stock_request.save()
        
        return Response({
            'success': True,
            'message': 'Stock request rejected'
        })
    
    @action(detail=False, methods=['get'])
    def pending(self, request):
        """Get all pending stock requests"""
        requests = self.get_queryset().filter(status='pending')
        serializer = self.get_serializer(requests, many=True)
        return Response(serializer.data)
